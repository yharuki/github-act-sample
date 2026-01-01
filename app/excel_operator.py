# coding: utf-8

__author__ = "Yuji Haruki (modifier) / Kohei, Watanabe <kohei.watanabe3@brother.co.jp> (original)"
__version__ = "2.1.0"
__date__ = "5 June 2024"

import sys
import os.path
from itertools import product
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.borders import BORDER_THIN, BORDER_THICK, BORDER_NONE
from warningMsgProvider import ExOpStatus, WarningMsgProvider
import shutil
import string
warning_msg_provider = WarningMsgProvider()


def col_num_to_excel_col_name(index):
    """
    インデックスをExcelの列名に変換する関数
    :param index: int, 変換するインデックス
    :return: str, Excelの列名
    """
    if index <= 0:
        raise ValueError("Index must be greater than zero.")
    letters = list(string.ascii_uppercase)
    num_letters = len(letters)

    column_name = ""
    while index > 0:
        remainder = (index - 1) % num_letters
        column_name = letters[remainder] + column_name
        index = (index - 1) // num_letters

    return column_name


def create_combined_col_params(target1, target2, loop_num_of_target2):
    combined_col_params = list(target1.values())
    for idx in range(loop_num_of_target2):
        for k, v in target2.items():
            combined_col_params.append(v)
    return combined_col_params


def write_test_specification(
    df: pd.DataFrame,
    sheet_name: str,
    summary: list,
    test_env_frame: list,
    writer: pd.ExcelWriter,
    config_excel: dict,
    merge_cells: bool,
) -> None:
    def is_test_intention_col(col_idx):
        col_num = col_idx + 1
        remainder = (col_num - len(config_excel["col_name"])) % len(
            config_excel["col_name_res_area"]
        )
        if col_num >= len(config_excel["col_name"]) and remainder == 0:
            return True
        else:
            return False

    test_env_frame_num = len(test_env_frame)

    df_excel = df.copy()
    df_excel.rename(columns=config_excel["col_name"], inplace=True)
    for idx in range(test_env_frame_num):
        for k, v in config_excel["col_name_res_area"].items():
            tmp_k = k + "_" + str(idx + 1)
            tmp_v = v + str(idx + 1)
            df_excel.rename(columns={tmp_k: tmp_v}, inplace=True)

    # マージできるようにマルチインデックス化
    # ※ マルチインデックス化の機能は、運用上のため将来的に削除したい
    items = [
        str(v)
        for k, v in config_excel["col_name"].items()
        if k in config_excel["index"] and config_excel["index"][k]
    ]
    df_excel.set_index(
        items,
        inplace=True,
    )

    # 出力する列の設定
    #  マルチインデックス化した列に加え、ここで指定した列を出力する
    output_cols = []
    for k, v in config_excel["col_name"].items():
        if k in config_excel["output"] and config_excel["output"][k]:
            output_cols.append(v)

    for idx in range(test_env_frame_num):
        for k, v in config_excel["col_name_res_area"].items():
            # 同じ列名が複数あると df_excel.to_excel でエラーとなるため
            # テスト環境枠が複数ある場合に備えて、当該領域は一時的に末尾にインデックスをつけた列名にしておく
            tmp_v = v + str(idx + 1)
            output_cols.append(tmp_v)

    # 書き出し開始行の設定
    #   概要 `summary` の行数に応じて、テスト項目表の開始位置を調整する
    summary_line_cnt = len(summary)
    if summary_line_cnt > config_excel["def_offset_row"]:
        tb_start_row = summary_line_cnt + 2
    else:
        tb_start_row = config_excel["def_offset_row"] + 2

    # データフレームをエクセルシートに変換
    df_excel.to_excel(
        writer,
        sheet_name=sheet_name,
        merge_cells=merge_cells,
        startrow=tb_start_row - 1,
        columns=output_cols,
        index=True,  # マルチインデックス化しない場合は `False` を設定
    )

    worksheet = writer.sheets[sheet_name]

    # 一時的にテスト環境枠列名の末尾にインデックスをつけた状態を元に戻す
    for i in range(test_env_frame_num):
        start_col = (
            len(config_excel["col_name"]) + len(config_excel["col_name_res_area"]) * i
        )
        for idx, key in enumerate(config_excel["col_name_res_area"]):
            col = col_num_to_excel_col_name(start_col + idx)
            __col_address = col + str(tb_start_row)
            worksheet[__col_address].value = config_excel["col_name_res_area"][key]

            # テスト環境枠の名称を設定
            if idx == 0:
                __col_address = col + str(tb_start_row - 1)
                worksheet[__col_address].value = test_env_frame[i]

    # 行固定
    worksheet.freeze_panes = "A" + str(tb_start_row + 1)

    # 概要行の書き出し
    for idx, one_line in enumerate(summary):
        worksheet["E" + str(idx + 1)].value = one_line

    # ここからExcelデータの見た目を整えていく

    # 合計列数取得
    total_col_count = len(config_excel["col_name"]) - 1
    total_col_count += len(config_excel["col_name_res_area"]) * test_env_frame_num

    # 列のカラーインデックス作成
    arr_color_index = create_combined_col_params(
        config_excel["header_color"],
        config_excel["header_color_res_area"],
        test_env_frame_num,
    )

    # ヘッダーのスタイル設定
    for col_idx in range(total_col_count):
        col = col_num_to_excel_col_name(col_idx + 1)
        __col_address = col + str(tb_start_row)
        worksheet[__col_address].font = Font(
            name=config_excel["font"], b=True, color="000000", size=9
        )
        worksheet[__col_address].alignment = Alignment(
            text_rotation=255, vertical="center", horizontal="center", wrap_text=True
        )
        worksheet[__col_address].fill = PatternFill(
            patternType="solid", fgColor=arr_color_index[col_idx]
        )
        worksheet[__col_address].border = Border(
            left=Side(
                # テスト仕様列群と結果列群の境界は太線
                style="medium" if is_test_intention_col(col_idx) else BORDER_THIN
            ),
            right=Side(style=BORDER_THIN),
            top=Side(style=BORDER_THIN),
            bottom=Side(style=BORDER_THIN),
        )
        worksheet[__col_address].value = worksheet[__col_address].value.rstrip()
        worksheet.row_dimensions[tb_start_row].height = config_excel["height"]["header"]

    # 列幅インデックス
    arr_width_index = create_combined_col_params(
        config_excel["width"], config_excel["width_res_area"], test_env_frame_num
    )
    for col_idx in range(total_col_count):
        col_name = col_num_to_excel_col_name(col_idx + 1)
        worksheet.column_dimensions[col_name].width = arr_width_index[col_idx]

    # データセルのスタイル調整

    lv = list(config_excel["index"])
    lv_color_fill_flag = [False] * len(config_excel["index"])
    is_lv_row = False

    # 行ループ
    for row_idx in range(len(df_excel)):
        # 背景の着色フラグ準備
        mark = df_excel.iloc[row_idx]["MARK"]
        is_lv_row = mark in lv
        if is_lv_row:
            lv_idx = lv.index(mark)
            lv_color_fill_flag[lv_idx] = True
            lv_color_fill_flag[lv_idx + 1 :] = [False] * (
                len(lv_color_fill_flag) - lv_idx - 1
            )
        fill_color = ""

        # 水平位置インデックス
        arr_horizontal_index = create_combined_col_params(
            config_excel["horizontal"],
            config_excel["horizontal_res_area"],
            test_env_frame_num,
        )
        # 垂直位置インデックス
        arr_vertical_index = create_combined_col_params(
            config_excel["vertical"],
            config_excel["vertical_res_area"],
            test_env_frame_num,
        )
        # 列ループ
        for col_idx in range(total_col_count):
            col_name = col_num_to_excel_col_name(col_idx + 1)
            __col_address = col_name + f"{row_idx + 1 + tb_start_row}"

            # 背景色の設定
            if col_idx < len(lv_color_fill_flag) and lv_color_fill_flag[col_idx]:
                fill_color = arr_color_index[col_idx]
            elif is_lv_row:
                if any(
                    lv_color_fill_flag[col_idx + 1 :]
                ):  # 通常はないがテスト観点レベルの追い越しがあった場合の対応  ex. lv1-lv2-lv4
                    fill_color = ""
                else:
                    pass
            else:
                fill_color = ""

            if fill_color:
                worksheet[__col_address].fill = PatternFill(
                    patternType="solid", fgColor=fill_color
                )

            # 罫線の設定
            border_style = {
                "left": BORDER_NONE,
                "right": BORDER_NONE,
                "top": BORDER_NONE,
                "bottom": BORDER_NONE,
            }

            # 先頭列
            if col_idx == 0:
                border_style["left"] = BORDER_THIN
            # 末尾列
            if col_idx == total_col_count - 1:
                border_style["right"] = BORDER_THIN
            # 末尾行
            if row_idx == len(df_excel) - 1:
                border_style["bottom"] = BORDER_THIN

            # テスト観点行
            wrap_text = True
            shrink_to_fit = False
            if is_lv_row:
                # テスト観点列の番号を必要に応じて縮小表示
                if col_idx < len(config_excel["index"]):
                    shrink_to_fit = True

                wrap_text = False
                if col_idx < len(lv_color_fill_flag) and lv_color_fill_flag[col_idx]:
                    border_style["left"] = BORDER_THIN
                    if any(
                        lv_color_fill_flag[col_idx + 1 :]
                    ):  # 通常はないがテスト観点レベルの追い越しがあった場合の対応  ex. lv1-lv2-lv4
                        border_style["right"] = BORDER_THIN
                    if col_idx == lv.index(mark):
                        border_style["top"] = BORDER_THIN
                elif col_idx > lv.index(mark):
                    border_style["top"] = BORDER_THIN
                    border_style["bottom"] = BORDER_THIN

            # 項目行
            else:
                if col_idx < len(lv_color_fill_flag):
                    if lv_color_fill_flag[col_idx]:
                        border_style["left"] = BORDER_THIN
                        border_style["right"] = BORDER_THIN
                else:
                    border_style["left"] = BORDER_THIN
                    border_style["bottom"] = BORDER_THIN

            # テスト仕様列群と結果列群の境界は太線
            if is_test_intention_col(col_idx):
                border_style["left"] = "medium"

            worksheet[__col_address].border = Border(
                left=Side(style=border_style["left"]),
                right=Side(style=border_style["right"]),
                top=Side(style=border_style["top"]),
                bottom=Side(style=border_style["bottom"]),
            )

            worksheet[__col_address].alignment = Alignment(
                horizontal=arr_horizontal_index[col_idx],
                vertical=arr_vertical_index[col_idx],
                wrap_text=wrap_text,
                shrink_to_fit=shrink_to_fit,
            )
            worksheet[__col_address].font = Font(
                name=config_excel["font"], color="000000", size=9
            )


def convert_df_to_excel(
    dfs: list[pd.DataFrame],
    sheet_names: list[str],
    product_categories: list[str],
    summaries: list[list],
    test_env_frames: list[list],
    config_excel: dict,
    input_path: str,
    output_fn: str = "TestSpec.xlsm",
    merge_cells: bool = True,
) -> None:
    """
    convert_md_to_df()により生成されたデータフレームをエクセルシートに変換します
    生成したシートを指定のエクセルファイルに追加します

    Args:
        dfs:                convert_md_to_df()により生成されたデータフレーム
        sheet_names          シート名
        product_categories  製品カテゴリー
        summares:           タイトル名、および概要欄の入力文章
        test_env_frames:    テスト環境枠
        config_excel:       設定
        input_path:         エクセルのテンプレファイル
        output_fn:          出力先のファイル
        merge_cells:        テスト観点のセルを結合するかどうか（非サポート）

    Returns:
        None
    """

    # テンプレートからエクセルファイルを複製
    try:
        warning_msg_provider.setTargetFP(output_fn)

        if os.path.exists(output_fn):
            print("\n保存先のファイルが既に存在します " + "(" + output_fn + ")")
            while True:
                user_input = input("→ 上書きしますか? (y/n): ").lower()
                if user_input == 'y':
                    print("")
                    break
                elif user_input == 'n':
                    print(output_fn + " の書き込みをスキップしました\n")
                    return 
                else:
                    print("→ 'y' または 'n' いずれかのキーを押してください")

        shutil.copy2(input_path, output_fn)
    except PermissionError:
        msg = warning_msg_provider.buildMsg(ExOpStatus.ERROR_CODE_1.value)
        print(msg)
        input("何かキーを押してください...")
        sys.exit(1)

    writer = pd.ExcelWriter(
        output_fn, mode="a", engine="openpyxl", engine_kwargs={"keep_vba": True}
    )

    # テスト項目シート追加
    try:
        for idx, df in enumerate(dfs):
            sheet_name = (
                sheet_names[idx] if sheet_names[idx] != "" else f"Sheet{str(idx + 1)}"
            )
            summary = summaries[idx]
            test_env_frame = test_env_frames[idx]
            write_test_specification(
                df,
                sheet_name,
                summary,
                test_env_frame,
                writer,
                config_excel,
                merge_cells,
            )

            # シート移動
            wb = writer.book
            ws = wb[sheet_name]
            wb.move_sheet(ws, offset=-3)

        # 製品カテゴリの表紙シート選定
        product_categorie = product_categories[0]
        summary_sheet_title = "表紙_共通"
        specified_summary_sheet_title = "表紙_" + product_categorie
        if specified_summary_sheet_title in [ws.title for ws in wb.worksheets]:
            summary_sheet_title = specified_summary_sheet_title
        for ws in wb.worksheets:
            if ws.title.startswith("表紙"):
                if ws.title != summary_sheet_title:
                    wb.remove(ws)
                else:
                    if summary_sheet_title == "表紙_共通":
                        ws.cell(row=1, column=1, value=product_categorie)
                    ws.title = '表紙'

        # シートのタブ選択状態を解除して先頭シートを選択
        for ws in wb.worksheets:
            ws.sheet_view.tabSelected = False
        wb.active = wb.worksheets[0]

        # 保存
        writer.close()
    except ValueError as e:
        msg = warning_msg_provider.buildMsg(ExOpStatus.ERROR_CODE_2.value)
        print(msg)
        input("何かキーを押してください...")
        sys.exit(1)

    # MEMO
    # Excelのアドインを Python から実行することも可能ではあるが、以下の理由から見送る
    # - 別途、アドインファイルの読み込みが必要で処理が煩雑になる
    # - 上記ファイルの格納場所は環境により異なる
    # - 読み込みに多少時間がかかる


def convert_excel_to_df(input_path: str) -> tuple[dict, str]:
    """
    Args:
        input_path:        入力ファイルパス

    Returns:
        df:                データフレーム型テスト項目書
    """

    IGNORED_SHEET_NAME = ["レビュー記録", "消化率", "マクロ起動"]

    dfs = {}
    product_category = "共通"
    wb = pd.ExcelFile(input_path)
    for sheet_name in wb.sheet_names:
        if sheet_name in IGNORED_SHEET_NAME:
            continue
        df = pd.read_excel(input_path, sheet_name=sheet_name, header=None, dtype=str)
        if sheet_name == "表紙":
            product_category = df[0][0] # セルA1
            continue
        df.fillna("", inplace=True)
        dfs[sheet_name] = df

    return dfs, product_category
